import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# Retrieve personal information from environment variables
NAME = os.getenv('NAME')
ADDRESS = os.getenv('ADDRESS')
CITY_STATE_ZIP = os.getenv('CITY_STATE_ZIP')
SOCIAL_HANDLE = os.getenv('SOCIAL_HANDLE')

# Default hourly rates for different types of audiobook work
DEFAULT_RATES = {
    "audiobook editing": 50.0,
    "audiobook proofing": 25.0,
    "extra editing": 50.0
}

def get_items(rates_dict):
    """
    Prompts the user to input the hours worked for each type of work listed in rates_dict.
    
    Args:
        rates_dict (dict): A dictionary with types of work as keys and their hourly rates as values.
        
    Returns:
        list: A list of dictionaries, each containing details of work type, hours, rate, and total price.
    """
    items = []
    for work_type in rates_dict.keys():
        performed = input(f"Did you perform: {work_type}? ").lower()
        if performed == "yes":
            try:
                hours = float(input("Enter hours: "))
                rate = rates_dict[work_type]
                total_price = hours * rate
                items.append({
                    "Type of Work": work_type,
                    "Hours": hours,
                    "Hourly Rate": rate,
                    "Total Price": total_price
                })
            except ValueError:
                print("Invalid input for hours. Please enter a numeric value.")
    return items

def get_invoice_details():
    """
    Prompts the user to input customer and invoice details and collects itemized work details.
    
    Returns:
        dict: A dictionary containing customer name, address, invoice date, book title, and itemized work details.
    """
    customer_name = input("Enter customer name: ")
    customer_address = input("Enter customer address: ")
    invoice_date = input("Enter invoice date (YYYY-MM-DD): ")
    book_title = input("Enter the title of the book: ")

    items_list = get_items(DEFAULT_RATES)

    return {
        "customer_name": customer_name,
        "customer_address": customer_address,
        "invoice_date": invoice_date,
        "book_title": book_title,
        "items": items_list
    }

def create_invoice(invoice_details):
    """
    Creates an invoice Excel file based on provided details.
    
    Args:
        invoice_details (dict): A dictionary containing customer name, address, invoice date, book title, and itemized work details.
    """
    customer_name = invoice_details["customer_name"]
    customer_address = invoice_details["customer_address"]
    invoice_date = invoice_details["invoice_date"]
    book_title = invoice_details["book_title"]
    items = invoice_details["items"]

    df = pd.DataFrame(items)
    total_amount = df["Total Price"].sum()

    # Generate a unique filename based on customer name and invoice date
    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"invoice_{customer_name.replace(' ', '_')}_{invoice_date}_{timestamp}.xlsx"

    # Create an Excel writer object
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        workbook = writer.book
        sheet = workbook.create_sheet("Invoice")

        # Add header information
        header_font = Font(size=14, bold=True)
        subheader_font = Font(size=12, bold=True)
        title_font = Font(size=16, bold=True)
        
        sheet["A1"] = NAME
        sheet["A1"].font = title_font
        sheet["A2"] = ADDRESS
        sheet["A3"] = CITY_STATE_ZIP
        sheet["A4"] = SOCIAL_HANDLE
        sheet["A6"] = "Bill To:"
        sheet["A6"].font = subheader_font
        sheet["B6"] = customer_name
        sheet["B7"] = customer_address

        sheet["D1"] = "INVOICE"
        sheet["D1"].font = title_font
        sheet["D3"] = "INVOICE NUMBER"
        sheet["E3"] = f"KG {datetime.now().strftime('%m-%d-%y')}"
        sheet["D4"] = "INVOICE DATE"
        sheet["E4"] = invoice_date

        # Add column headers for the itemized section
        itemized_header = ["Date", "Description", "Total Run Time", "Rate", "Total"]
        for col_num, header in enumerate(itemized_header, 1):
            cell = sheet.cell(row=9, column=col_num)
            cell.value = header
            cell.font = subheader_font
            cell.alignment = Alignment(horizontal="center")

        # Add the book title above the description column
        sheet.cell(row=10, column=2, value=book_title)
        sheet.cell(row=10, column=2).font = subheader_font
        sheet.cell(row=10, column=2).alignment = Alignment(horizontal="center")

        # Add the itemized details
        row_num = 11
        for item in items:
            sheet.cell(row=row_num, column=1, value=invoice_date)
            sheet.cell(row=row_num, column=2, value=item["Type of Work"].title())
            sheet.cell(row=row_num, column=3, value=f'{int(item["Hours"]):02}:{int((item["Hours"] * 60) % 60):02}:{int((item["Hours"] * 3600) % 60):02}')
            sheet.cell(row=row_num, column=4, value=f"${item['Hourly Rate']:.2f}")
            sheet.cell(row=row_num, column=5, value=f"${item['Total Price']:.2f}")
            row_num += 1

        # Add total amount at the end
        sheet.cell(row=row_num + 1, column=4, value="TOTAL")
        sheet.cell(row=row_num + 1, column=5, value=f"${total_amount:.2f}")
        sheet.cell(row=row_num + 2, column=5, value="Please Pay")
        sheet.cell(row=row_num + 4, column=2, value="Make All Checks Payable To:")
        sheet.cell(row=row_num + 5, column=2, value=NAME)
        sheet.cell(row=row_num + 6, column=2, value=ADDRESS)
        sheet.cell(row=row_num + 7, column=2, value=CITY_STATE_ZIP)

        # Adjust column widths for better readability
        for col in sheet.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column].width = max_length + 2

    print(f"Invoice created and saved as '{filename}'")

def main():
    """
    Main function to gather invoice details and create an invoice.
    """
    invoice_details = get_invoice_details()
    create_invoice(invoice_details)

if __name__ == "__main__":
    main()
